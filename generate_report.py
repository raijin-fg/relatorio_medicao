"""
CPTM - Gerador de Relatório Mensal de TI
Lê dados do banco SQLite e gera relatório Excel em output/
"""

import sqlite3
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from isodate import parse_duration
import isodate

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DB_PATH = Path(__file__).parent / "data" / "prd_data.db"
OUTPUT_DIR = Path(__file__).parent / "output"
NUM_DRIVERS = 6
MAX_NETAPP_HOURS = 10.0   # duração alvo máxima por política NetApp (horas)

# Cores
C_HEADER_BG   = "1F3864"   # azul escuro
C_HEADER_FT   = "FFFFFF"   # branco
C_SECTION_BG  = "D6E4F0"   # azul claro
C_SECTION_FT  = "1F3864"   # azul escuro
C_ALT_ROW     = "F2F7FC"   # azul muito claro
C_GREEN_BG    = "C6EFCE"   # verde claro
C_GREEN_FT    = "276221"   # verde escuro
C_YELLOW_BG   = "FFEB9C"   # amarelo claro
C_YELLOW_FT   = "9C6500"   # amarelo escuro
C_RED_BG      = "FFC7CE"   # vermelho claro
C_RED_FT      = "9C0006"   # vermelho escuro
C_ORANGE_BG   = "FFD966"   # laranja claro
C_ORANGE_FT   = "7D4E00"   # Laranja
C_GRAY_BG     = "D9D9D9"   # Cinza
C_WHITE       = "FFFFFF"   # branco
C_ACCENT      = "2E75B6"   # 

FONT_NAME = "Arial"

# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, values, widths=None, bg=C_HEADER_BG, ft=C_HEADER_FT, size=10):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _font(bold=True, size=size, color=ft)
        cell.fill = _fill(bg)
        cell.alignment = _align("center")
        cell.border = _border()
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

def _data_row(ws, row, values, alt=False):
    bg = C_ALT_ROW if alt else C_WHITE
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _font()
        cell.fill = _fill(bg)
        cell.alignment = _align()
        cell.border = _border()

def _section_title(ws, row, text, ncols, bg=C_SECTION_BG, ft=C_SECTION_FT):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _font(bold=True, size=11, color=ft)
    cell.fill = _fill(bg)
    cell.alignment = _align("left")
    cell.border = _border()

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _iso_to_hours(iso_str):
    """Converte string ISO 8601 duration (PT13H47M2S) para horas float."""
    try:
        return isodate.parse_duration(iso_str).total_seconds() / 3600
    except Exception:
        return 0.0

def _iso_to_hhmm(iso_str):
    """Converte ISO duration para string HH:MM legível."""
    try:
        total_sec = int(isodate.parse_duration(iso_str).total_seconds())
        h, rem = divmod(total_sec, 3600)
        m = rem // 60
        return f"{h:02d}:{m:02d}"
    except Exception:
        return "00:00"

def _parse_dt(iso_str):
    """Parse ISO 8601 datetime para datetime ciente de UTC."""
    if not iso_str:
        return None
    return datetime.fromisoformat(iso_str.replace("Z", "+00:00"))

def _kb_to_tb(kb):
    return kb / (1024 ** 3)

def _kb_to_gb(kb):
    return kb / (1024 ** 2)

def _sla_color(pct):
    if pct is None:
        return C_GRAY_BG, "000000"
    if pct >= 99.9:
        return C_GREEN_BG, C_GREEN_FT
    if pct >= 99.0:
        return C_YELLOW_BG, C_YELLOW_FT
    return C_RED_BG, C_RED_FT

# ---------------------------------------------------------------------------
# Leitura do banco
# ---------------------------------------------------------------------------

def _load(db_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        cursor = conn.cursor()


        avail   = conn.execute("SELECT * FROM zabbix_availability ORDER BY group_name, host").fetchall()
        jobs    = conn.execute("SELECT * FROM netbackup_jobs ORDER BY start_time").fetchall()
        pols    = conn.execute("SELECT * FROM netbackup_policies").fetchall()
        tapes   = conn.execute("SELECT * FROM netbackup_tapes ORDER BY snapshot_date").fetchall()
        volumes = conn.execute("""
            SELECT * FROM netapp_volumes
            ORDER BY snapshot_date, volume_name
        """).fetchall()
        tapes_history = conn.execute("SELECT * FROM netbackup_tapes_history ORDER BY snapshot_month").fetchall()

        return avail, jobs, pols, tapes, volumes, tapes_history
    finally:
        conn.close()

# ---------------------------------------------------------------------------
# Aba 1 — Capa / Resumo Executivo
# ---------------------------------------------------------------------------

def _sheet_capa(wb, avail, jobs, volumes, report_month):
    ws = wb.create_sheet("Capa")
    ws.sheet_view.showGridLines = False

    # Título principal
    ws.merge_cells("A1:H2")
    c = ws["A1"]
    c.value = "RELATÓRIO MENSAL DE TI — CPTM"
    c.font = _font(bold=True, size=18, color=C_HEADER_FT)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _align("center")

    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value = f"Competência: {report_month}"
    c.font = _font(bold=True, size=12, color=C_SECTION_FT)
    c.fill = _fill(C_SECTION_BG)
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    # ---- KPIs de disponibilidade por grupo ----
    row = 5
    _section_title(ws, row, "DISPONIBILIDADE — RESUMO POR GRUPO", 8)

    row += 1
    _header_row(ws, row, ["Grupo", "Total Hosts", "≥ 99,9%", "≥ 99,0%", "< 99,0%", "Offline (0%)", "Uptime Médio", "SLA Status"])

    groups = {}
    for r in avail:
        g = r["group_name"]
        pct = r["availability_pct"]
        if isinstance(pct, str):
            pct = None
        if g not in groups:
            groups[g] = []
        groups[g].append(pct)

    row += 1
    for g, vals in sorted(groups.items()):
        numeric = [v for v in vals if v is not None]
        total   = len(vals)
        above999 = sum(1 for v in numeric if v >= 99.9)
        above99  = sum(1 for v in numeric if 99.0 <= v < 99.9)
        below99  = sum(1 for v in numeric if v < 99.0)
        offline  = sum(1 for v in numeric if v == 0.0)
        avg      = sum(numeric) / len(numeric) if numeric else None

        bg, ft = _sla_color(avg)
        sla_txt = "✔ OK" if avg and avg >= 99.9 else ("⚠ ATENÇÃO" if avg and avg >= 99.0 else "✖ CRÍTICO")

        for col, val in enumerate([g, total, above999, above99, below99, offline,
                                    f"{avg:.4f}%" if avg is not None else "N/A", sla_txt], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(bold=(col == 8), color=ft if col in (7, 8) else "000000")
            cell.fill = _fill(bg if col in (7, 8) else C_WHITE)
            cell.alignment = _align("center")
            cell.border = _border()
        row += 1

    # ---- KPIs de backup ----
    row += 1
    _section_title(ws, row, "BACKUP — RESUMO DO MÊS", 8)

    row += 1
    _header_row(ws, row, ["Métrica", "Valor"])

    total_jobs  = len(jobs)
    total_tb    = sum(_kb_to_tb(j["kilobytes"]) for j in jobs)
    netapp_jobs = [j for j in jobs if j["job_policy"].startswith("NETAPP")]
    vmware_jobs = [j for j in jobs if j["job_policy"].startswith("VMWARES")]
    oracle_jobs = [j for j in jobs if "ORACLE" in j["job_policy"]]
    exch_jobs   = [j for j in jobs if "EXCHANGE" in j["job_policy"]]

    metrics = [
        ("Total de jobs executados no mês", total_jobs),
        ("Volume total transferido (TB)", f"{total_tb:.2f}"),
        ("Volume NetApp (TB)", f"{sum(_kb_to_tb(j['kilobytes']) for j in netapp_jobs):.2f}"),
        ("Volume VMware (TB)", f"{sum(_kb_to_tb(j['kilobytes']) for j in vmware_jobs):.2f}"),
        ("Volume Oracle (TB)", f"{sum(_kb_to_tb(j['kilobytes']) for j in oracle_jobs):.2f}"),
        ("Volume Exchange (TB)", f"{sum(_kb_to_tb(j['kilobytes']) for j in exch_jobs):.2f}"),
        ("Job mais longo", max((_iso_to_hhmm(j["elapsed_time"]) for j in jobs), default="—")),
        ("Política mais longa",  max(jobs, key=lambda j: _iso_to_hours(j["elapsed_time"]))["job_policy"] if jobs else "—"),
    ]

    row += 1
    for i, (label, val) in enumerate(metrics):
        bg = C_ALT_ROW if i % 2 else C_WHITE
        for col, v in enumerate([label, val], 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = _font(bold=(col == 1))
            cell.fill = _fill(bg)
            cell.alignment = _align()
            cell.border = _border()
        row += 1

    # ---- NetApp storage ----
    row += 1
    _section_title(ws, row, "ARMAZENAMENTO NETAPP — SNAPSHOT MAIS RECENTE", 8)
    row += 1
    _header_row(ws, row, ["Métrica", "Valor"])

    latest = max((v["snapshot_date"] for v in volumes), default=None)
    latest_vols = [v for v in volumes if v["snapshot_date"] == latest] if latest else []
    total_alloc = sum(v["total_gb"] for v in latest_vols)
    total_used  = sum(v["used_gb"]  for v in latest_vols)
    total_free  = sum(v["free_gb"]  for v in latest_vols)
    pct_used    = (total_used / total_alloc * 100) if total_alloc else 0

    stor_metrics = [
        ("Data do snapshot", latest or "—"),
        ("Capacidade total alocada (TB)", f"{total_alloc/1024:.2f}"),
        ("Utilizado (TB)",  f"{total_used/1024:.2f}"),
        ("Disponível (TB)", f"{total_free/1024:.2f}"),
        ("% Utilizado", f"{pct_used:.1f}%"),
        ("Volumes monitorados", len(latest_vols)),
    ]
    row += 1
    for i, (label, val) in enumerate(stor_metrics):
        bg = C_ALT_ROW if i % 2 else C_WHITE
        for col, v in enumerate([label, val], 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = _font(bold=(col == 1))
            cell.fill = _fill(bg)
            cell.alignment = _align()
            cell.border = _border()
        row += 1

    _set_col_widths(ws, [38, 14, 10, 10, 10, 12, 14, 14])

# ---------------------------------------------------------------------------
# Aba 2 — Disponibilidade (icmpping)
# ---------------------------------------------------------------------------

def _sheet_disponibilidade(wb, avail):
    ws = wb.create_sheet("Disponibilidade")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    _header_row(ws, 1,
        ["Grupo", "Host", "Disponibilidade (%)", "SLA", "Observação"],
        widths=[28, 55, 22, 14, 30])

    # Ordenar: grupo asc, disponibilidade asc (piores primeiro dentro do grupo)
    rows_sorted = sorted(avail, key=lambda r: (
        r["group_name"],
        r["availability_pct"] if isinstance(r["availability_pct"], float) else -1
    ))

    data_start = 2
    for i, r in enumerate(rows_sorted):
        row = data_start + i
        pct = r["availability_pct"]
        pct_num = pct if isinstance(pct, float) else None
        pct_str = f"{pct_num:.4f}%" if pct_num is not None else str(pct)

        bg, ft = _sla_color(pct_num)

        if pct_num is None:
            sla = "Sem dados"
            obs = "Item não retornou valor numérico"
        elif pct_num == 0.0:
            sla = "OFFLINE"
            obs = "Host indisponível durante todo o período"
        elif pct_num >= 99.9:
            sla = "✔ OK"
            obs = ""
        elif pct_num >= 99.0:
            sla = "⚠ ATENÇÃO"
            obs = f"Downtime ≈ {(100 - pct_num) / 100 * 43200:.1f} min no mês"
        else:
            sla = "✖ CRÍTICO"
            obs = f"Downtime ≈ {(100 - pct_num) / 100 * 43200:.0f} min no mês"

        alt = i % 2 == 1
        row_bg = C_ALT_ROW if alt else C_WHITE

        vals = [r["group_name"], r["host"], pct_str, sla, obs]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col in (3, 4):
                cell.font = _font(bold=(col == 4), color=ft)
                cell.fill = _fill(bg)
            else:
                cell.font = _font()
                cell.fill = _fill(row_bg)
            cell.alignment = _align("center" if col in (3, 4) else "left")
            cell.border = _border()

    # Resumo por grupo no final
    data_end = data_start + len(rows_sorted)
    row = data_end + 1
    _section_title(ws, row, "RESUMO POR GRUPO", 5)
    row += 1
    _header_row(ws, row, ["Grupo", "Hosts", "Uptime Médio", "Piores hosts (< 99%)", ""])

    groups = {}
    for r in avail:
        g = r["group_name"]
        pct = r["availability_pct"]
        if g not in groups:
            groups[g] = []
        if isinstance(pct, float):
            groups[g].append((r["host"], pct))

    row += 1
    for g, items in sorted(groups.items()):
        vals_only = [v for _, v in items]
        avg = sum(vals_only) / len(vals_only) if vals_only else None
        worst = [h for h, v in sorted(items, key=lambda x: x[1])[:3] if v < 99.0]
        worst_str = "; ".join(worst) if worst else "—"
        bg, ft = _sla_color(avg)
        for col, val in enumerate([g, len(items), f"{avg:.4f}%" if avg else "N/A", worst_str, ""], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(bold=(col in (1, 3)), color=ft if col == 3 else "000000")
            cell.fill = _fill(bg if col == 3 else C_WHITE)
            cell.alignment = _align("center" if col in (2, 3) else "left")
            cell.border = _border()
        row += 1

# ---------------------------------------------------------------------------
# Aba 3 — Backup Geral
# ---------------------------------------------------------------------------

def _sheet_backup_geral(wb, jobs):
    ws = wb.create_sheet("Backup — Visão Geral")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    headers = ["Política", "Schedule", "Início", "Término", "Duração",
               "Volume (GB)", "Volume (TB)", "Velocidade (MB/s)", "Categoria"]
    widths  = [28, 22, 22, 22, 12, 14, 12, 18, 18]
    _header_row(ws, 1, headers, widths)

    def _categoria(policy):
        p = policy.upper()
        if p.startswith("NETAPP"):   return "NetApp"
        if p.startswith("VMWARES"): return "VMware"
        if "ORACLE" in p:           return "Oracle"
        if "EXCHANGE" in p:         return "Exchange"
        if p.startswith("VMWARE"):  return "VMware"
        return "Windows/Outros"

    jobs_sorted = sorted(jobs, key=lambda j: j["start_time"] or "")

    for i, j in enumerate(jobs_sorted):
        row = 2 + i
        alt = i % 2 == 1
        start = _parse_dt(j["start_time"])
        end   = _parse_dt(j["end_time"])
        gb    = _kb_to_gb(j["kilobytes"])
        tb    = _kb_to_tb(j["kilobytes"])
        mbps  = j["kb_sec"] / 1024 if j["kb_sec"] else 0

        vals = [
            j["job_policy"],
            j["job_schedule"],
            start.astimezone().strftime("%d/%m/%Y %H:%M") if start else "",
            end.astimezone().strftime("%d/%m/%Y %H:%M") if end else "",
            _iso_to_hhmm(j["elapsed_time"]),
            round(gb, 2),
            round(tb, 4),
            round(mbps, 1),
            _categoria(j["job_policy"]),
        ]
        _data_row(ws, row, vals, alt)

    # Totais por categoria
    data_end = 1 + len(jobs_sorted)
    row = data_end + 2
    _section_title(ws, row, "TOTAIS POR CATEGORIA", len(headers))
    row += 1
    _header_row(ws, row, ["Categoria", "Qtd Jobs", "Volume Total (TB)", "Duração Total (h)", "Velocidade Média (MB/s)"])

    cats = {}
    for j in jobs:
        cat = _categoria(j["job_policy"])
        cats.setdefault(cat, []).append(j)

    row += 1
    for cat, jlist in sorted(cats.items()):
        total_tb  = sum(_kb_to_tb(j["kilobytes"]) for j in jlist)
        total_h   = sum(_iso_to_hours(j["elapsed_time"]) for j in jlist)
        avg_mbps  = sum((j["kb_sec"] or 0) / 1024 for j in jlist) / len(jlist)
        for col, val in enumerate([cat, len(jlist), round(total_tb, 2),
                                    round(total_h, 1), round(avg_mbps, 1)], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(bold=(col == 1))
            cell.fill = _fill(C_ALT_ROW if row % 2 else C_WHITE)
            cell.alignment = _align("center" if col > 1 else "left")
            cell.border = _border()
        row += 1

    # Top 10 jobs mais longos
    row += 1
    _section_title(ws, row, "TOP 10 — JOBS MAIS LONGOS", len(headers))
    row += 1
    _header_row(ws, row, ["Política", "Duração", "Volume (TB)", "Velocidade (MB/s)", "Início"])

    top10 = sorted(jobs, key=lambda j: _iso_to_hours(j["elapsed_time"]), reverse=True)[:10]
    row += 1
    for i, j in enumerate(top10):
        alt = i % 2 == 1
        start = _parse_dt(j["start_time"])
        for col, val in enumerate([
            j["job_policy"],
            _iso_to_hhmm(j["elapsed_time"]),
            round(_kb_to_tb(j["kilobytes"]), 4),
            round((j["kb_sec"] or 0) / 1024, 1),
            start.astimezone().strftime("%d/%m/%Y %H:%M") if start else "",
        ], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font()
            cell.fill = _fill(C_YELLOW_BG if i == 0 else (C_ALT_ROW if alt else C_WHITE))
            cell.alignment = _align("center" if col > 1 else "left")
            cell.border = _border()
        row += 1

# ---------------------------------------------------------------------------
# Aba 4 — Otimização de Janelas (Gantt + análise de drivers)
# ---------------------------------------------------------------------------

def _sheet_janelas(wb, jobs):
    """
    Aba 4 — Otimização de Janelas de Backup (Fim de Semana de 63h com 6 Drivers)
    Analisa a concorrência real e sugere o melhor horário de início dentro da janela.
    """
    ws = wb.create_sheet("Otimização — Janelas")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Título principal da aba
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "ANÁLISE DE CONCORRÊNCIA E REESCALONAMENTO DE JANELAS (FULL WEEKEND)"
    title_cell.font = _font(bold=True, size=11, color=C_HEADER_FT)
    title_cell.fill = _fill(C_HEADER_BG)
    title_cell.alignment = _align("left", "center")
    ws.row_dimensions[1].height = 26

    # Cabeçalhos exatos solicitados
    headers = [
        "Job ID", "Politica", "Início", "Término", "Duração (h)", 
        "Volume (TB)", "Concorrentes no Pico", "Conflito de Drivers?", "Sugestão de Início"
    ]
    _header_row(ws, 2, headers, widths=[12, 28, 18, 18, 14, 14, 22, 20, 45])
    ws.row_dimensions[2].height = 22

    # 1. MAPEAMENTO DE CONCORRÊNCIA POR HORA (Mapeia o mês em timestamps de hora em hora)
    hourly_concurrency = {}
    fds_jobs = []
    
    for j in jobs:
        try:
            job_dict = dict(j)
        except:
            job_dict = j

        # Coleta estritamente das colunas reais existentes no banco
        job_id = job_dict.get("job_id")
        job_policy = job_dict.get("job_policy")
        start_time = job_dict.get("start_time")
        end_time = job_dict.get("end_time")
        elapsed_time = job_dict.get("elapsed_time")
        kilobytes = job_dict.get("kilobytes")

        if not start_time or not end_time:
            continue
            
        dt_start = _parse_dt(start_time)
        dt_end = _parse_dt(end_time)
        
        dur_hours = _iso_to_hours(elapsed_time)
        dur_sec = dur_hours * 3600
        tb = _kb_to_tb(kilobytes) if kilobytes else 0.0
        
        # Filtro básico: foca em jobs que realmente têm impacto na janela (mais de 10 min ou volumosos)
        if dur_sec < 600 and tb < 0.05: 
            continue
            
        fds_jobs.append({
            "id": job_id,
            "policy": job_policy,
            "start": dt_start,
            "end": dt_end,
            "hours": dur_hours,
            "tb": tb
        })
        
        # Popula a linha do tempo de concorrência hora a hora
        start_ts = int(dt_start.timestamp() // 3600) * 3600
        end_ts = int(dt_end.timestamp() // 3600) * 3600
        for ts in range(start_ts, end_ts + 3600, 3600):
            hourly_concurrency[ts] = hourly_concurrency.get(ts, 0) + 1

    # 2. CALCULAR PICO, CONFLITO E BUSCAR JANELA IDEAL (Sexta 18h às Segunda 09h)
    row_idx = 3
    for job in fds_jobs:
        dt_start = job["start"]
        dt_end = job["end"]
        
        # Determina o Pico de concorrentes que estavam ativos ao mesmo tempo que este job
        job_start_ts = int(dt_start.timestamp() // 3600) * 3600
        job_end_ts = int(dt_end.timestamp() // 3600) * 3600
        
        concorrentes_no_pico = 0
        for ts in range(job_start_ts, job_end_ts + 3600, 3600):
            cnt = hourly_concurrency.get(ts, 0)
            if cnt > concorrentes_no_pico:
                concorrentes_no_pico = cnt
                
        # Se o pico de jobs rodando juntos passou do limite de 6 drivers da CPTM
        conflito = "Sim" if concorrentes_no_pico > NUM_DRIVERS else "Não"
        
        # --- ALGORITMO DE SUGESTÃO DE INÍCIO OTIMIZADO ---
        # Vamos simular o início deste backup em cada hora do Fim de Semana (FDS) daquela mesma semana
        # Janela Alvo: Começa Sexta às 18:00 e precisa terminar até Segunda às 09:00
        
        # Encontra a Sexta-feira daquela semana correspondente ao job
        # weekday: 0=Seg, 1=Ter, 2=Qua, 3=Qui, 4=Sex, 5=Sáb, 6=Dom
        dias_para_sexta = 4 - dt_start.weekday()
        sexta_referencia = dt_start + pd.Timedelta(days=dias_para_sexta) # type: ignore
        
        base_sexta_18h = datetime(
            sexta_referencia.year, sexta_referencia.month, sexta_referencia.day, 
            18, 0, 0, tzinfo=timezone.utc
        )
        limite_segunda_09h = base_sexta_18h + pd.Timedelta(hours=63) # type: ignore (63 horas depois = Segunda 09h)
        
        melhor_inicio = None
        menor_carga_concorrencia = float("inf")
        
        # Duração do job em horas inteiras arredondada para cima para segurança de simulação
        job_hours_needed = max(1, int(job["hours"] + 0.5))
        
        # Testamos cada janela inicial de hora em hora da Sexta 18h até Segunda 09h
        for offset_horas in range(63):
            sim_start = base_sexta_18h + pd.Timedelta(hours=offset_horas) # type: ignore
            sim_end = sim_start + pd.Timedelta(hours=job_hours_needed) # type: ignore
            
            # Validação Crítica: O job simulado consegue terminar antes de Segunda às 09:00?
            if sim_end > limite_segunda_09h:
                break # Daqui para a frente nenhum outro horário respeitará o limite de segunda
                
            # Calcula a soma de concorrência que esse bloco enfrentaria
            carga_concorrencia_janela = 0
            sim_start_ts = int(sim_start.timestamp() // 3600) * 3600
            for h_ts in range(sim_start_ts, sim_start_ts + (job_hours_needed * 3600), 3600):
                carga_concorrencia_janela += hourly_concurrency.get(h_ts, 0)
                
            # Queremos o horário que tenha a menor média de concorrência acumulada nos drivers
            if carga_concorrencia_janela < menor_carga_concorrencia:
                menor_carga_concorrencia = carga_concorrencia_janela
                melhor_inicio = sim_start
                
        # Define a mensagem final da sugestão
        if conflito == "Não":
            sugestao = "✔ Manter horário atual (Sem saturação de drivers)."
            cell_fill = None
            cell_font_color = "000000"
        elif melhor_inicio:
            # Tradução amigável do dia da semana sugerido
            dias_traducao = {4: "Sex", 5: "Sáb", 6: "Dom", 0: "Seg"}
            dia_str = dias_traducao.get(melhor_inicio.weekday(), "FDS")
            sugestao = f"Mover para {dia_str} às {melhor_inicio.strftime('%H:%M')} (Janela mais ociosa)."
            cell_fill = C_YELLOW_BG
            cell_font_color = C_YELLOW_FT
        else:
            sugestao = "⚠ Janela de FDS cheia. Recomendado Split de política."
            cell_fill = C_RED_BG
            cell_font_color = C_RED_FT

        # Escrita dos dados na planilha seguindo a formatação padrão do projeto
        ws.cell(row=row_idx, column=1, value=job["id"])
        ws.cell(row=row_idx, column=2, value=job["policy"])
        ws.cell(row=row_idx, column=3, value=dt_start.astimezone().strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row_idx, column=4, value=dt_end.astimezone().strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row_idx, column=5, value=round(job["hours"], 2))
        ws.cell(row=row_idx, column=6, value=job["tb"])
        ws.cell(row=row_idx, column=7, value=f"{concorrentes_no_pico} Jobs ativos")
        ws.cell(row=row_idx, column=8, value=conflito)
        
        cell_sugestao = ws.cell(row=row_idx, column=9, value=sugestao)
        if cell_fill:
            cell_sugestao.fill = _fill(cell_fill)
            cell_sugestao.font = _font(bold=True, color=cell_font_color)

        # Estilização das linhas (Zebra e Bordas) mantendo a harmonia visual corporativa
        bg_row = C_ALT_ROW if row_idx % 2 == 0 else C_WHITE
        for col in range(1, 10):
            c = ws.cell(row=row_idx, column=col)
            if col != 9:  # Preserva o destaque de cor na coluna de sugestão
                c.fill = _fill(bg_row)
                c.font = _font()
            c.border = _border()
            if col in [1, 3, 4, 5, 6, 7, 8]:
                c.alignment = _align("center")
            else:
                c.alignment = _align("left")

        # Formatações de número do Excel
        ws.cell(row=row_idx, column=5).number_format = "#,##0.00' h'"
        ws.cell(row=row_idx, column=6).number_format = "#,##0.0000' TB'"
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Ajuste automático de largura por coluna (exceto a I que é larga por causa do texto longo)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter != 'I':
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# ---------------------------------------------------------------------------
# Aba 5 — Otimização NetApp (split de políticas)
# ---------------------------------------------------------------------------

def _parse_netapp_volumes_from_policy(backup_selections):
    """Extrai nomes de volumes de backup_selections."""
    if not backup_selections:
        return []
    volumes = []
    for part in backup_selections.split(";"):
        part = part.strip()
        if not part or part.upper().startswith("SET"):
            continue
        # Remove prefixo /cptm_nas/ e barra final
        name = re.sub(r"^/?cptm_nas/?", "", part, flags=re.IGNORECASE).strip("/")
        if name:
            volumes.append(name.upper())
    return volumes

def _sheet_netapp_split(wb, jobs, pols, volumes):
    ws = wb.create_sheet("Otimização — NetApp Split")
    ws.sheet_view.showGridLines = False

    # Mapa volume -> tamanho (último snapshot)
    latest_date = max((v["snapshot_date"] for v in volumes), default=None)
    vol_size = {}
    for v in volumes:
        if v["snapshot_date"] == latest_date:
            vol_size[v["volume_name"].upper()] = v["used_gb"]

    # Mapa de políticas NetApp e seus volumes
    netapp_pols = {}
    for p in pols:
        if p["policy_name"].upper().startswith("NETAPP") and p["backup_selections"]:
            vols = _parse_netapp_volumes_from_policy(p["backup_selections"])
            if vols:
                netapp_pols[p["policy_name"]] = vols

    # Jobs NetApp para calcular velocidade média real
    netapp_jobs = [j for j in jobs if j["job_policy"].upper().startswith("NETAPP")]
    avg_kbsec = (sum(j["kb_sec"] for j in netapp_jobs if j["kb_sec"]) /
                 max(1, sum(1 for j in netapp_jobs if j["kb_sec"])))
    avg_gbsec = avg_kbsec / (1024 ** 2)  # KB/s -> GB/s

    # ---- Seção 1: Mapa atual de volumes por política ----
    _section_title(ws, 1, "MAPA ATUAL — VOLUMES POR POLÍTICA NETAPP", 8)
    _header_row(ws, 2,
        ["Política", "Volume", "Tamanho Usado (GB)", "% da política",
         "Duração estimada (h)", "Duração real último job (h)", "Status", ""],
        widths=[18, 22, 22, 16, 22, 26, 14, 14])

    row = 3
    policy_summaries = {}
    for pol_name in sorted(netapp_pols):
        vols = netapp_pols[pol_name]
        total_gb_pol = sum(vol_size.get(v, 0) for v in vols)

        # Duração real do último job dessa política
        pol_jobs = [j for j in jobs if j["job_policy"] == pol_name]
        real_h = max((_iso_to_hours(j["elapsed_time"]) for j in pol_jobs), default=None)

        policy_summaries[pol_name] = {
            "volumes": vols,
            "total_gb": total_gb_pol,
            "real_h": real_h,
            "est_h": total_gb_pol / (avg_gbsec * 3600) if avg_gbsec > 0 else 0,
        }

        # Linha de cabeçalho da política
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1,
            value=f"  {pol_name}  —  {len(vols)} volumes  |  {total_gb_pol:.1f} GB usado  |  Duração real: {_iso_to_hhmm(pol_jobs[-1]['elapsed_time']) if pol_jobs else '—'}")
        cell.font = _font(bold=True, size=10, color=C_SECTION_FT)
        cell.fill = _fill(C_SECTION_BG)
        cell.border = _border()
        row += 1

        for i, vol in enumerate(vols):
            gb = vol_size.get(vol, None)
            pct = (gb / total_gb_pol * 100) if (gb and total_gb_pol) else None
            est_h = (gb / (avg_gbsec * 3600)) if (gb and avg_gbsec > 0) else None
            alt = i % 2 == 1
            status = "✔ No banco" if gb else "⚠ Não encontrado"
            for col, val in enumerate([
                pol_name, vol,
                round(gb, 1) if gb else "—",
                f"{pct:.1f}%" if pct else "—",
                round(est_h, 2) if est_h else "—",
                round(real_h, 2) if real_h else "—",
                status, "",
            ], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = _font(color=(C_RED_FT if status.startswith("⚠") and col == 7 else "000000"))
                cell.fill = _fill(C_ALT_ROW if alt else C_WHITE)
                cell.alignment = _align("center" if col > 2 else "left")
                cell.border = _border()
            row += 1

    # ---- Seção 2: Sugestão de split ----
    row += 1
    _section_title(ws, row, f"SUGESTÃO DE SPLIT — META: MÁXIMO {MAX_NETAPP_HOURS}h POR POLÍTICA", 8)
    row += 1
    _header_row(ws, row,
        ["Nova política sugerida", "Volumes", "Total GB", "Duração estimada (h)",
         "Duração atual (h)", "Diferença (h)", "Recomendação", ""])
    row += 1

    # Coleta todos os volumes de todas as políticas NetApp com seus tamanhos
    all_vol_entries = []
    for pol_name, info in policy_summaries.items():
        for vol in info["volumes"]:
            gb = vol_size.get(vol, 0)
            all_vol_entries.append({"vol": vol, "gb": gb, "orig_pol": pol_name})

    # Bin packing greedy: ordena por tamanho desc, aloca em bins até MAX_NETAPP_HOURS
    all_vol_entries.sort(key=lambda x: x["gb"], reverse=True)
    bins = []  # lista de {"volumes": [], "total_gb": 0}

    max_gb_per_bin = MAX_NETAPP_HOURS * avg_gbsec * 3600 if avg_gbsec > 0 else float("inf")

    for entry in all_vol_entries:
        placed = False
        for b in bins:
            if b["total_gb"] + entry["gb"] <= max_gb_per_bin:
                b["volumes"].append(entry["vol"])
                b["total_gb"] += entry["gb"]
                placed = True
                break
        if not placed:
            bins.append({"volumes": [entry["vol"]], "total_gb": entry["gb"]})

    for i, b in enumerate(bins):
        pol_label = f"NETAPP_{chr(65 + i)}"  # NETAPP_A, NETAPP_B, ...
        est_h = b["total_gb"] / (avg_gbsec * 3600) if avg_gbsec > 0 else 0
        vols_str = ", ".join(b["volumes"])

        # Duração atual aproximada (média das políticas originais envolvidas)
        orig_pols = set(next((e["orig_pol"] for e in all_vol_entries if e["vol"] == v), "") for v in b["volumes"])
        current_h = max((policy_summaries[p]["real_h"] or 0 for p in orig_pols if p in policy_summaries), default=0)
        diff = est_h - current_h if current_h > 0 else None
        rec = "✔ Dentro do limite" if est_h <= MAX_NETAPP_HOURS else "⚠ Ainda excede limite — reduzir volumes"

        alt = i % 2 == 1
        for col, val in enumerate([
            pol_label, vols_str, round(b["total_gb"], 1), round(est_h, 2),
            round(current_h, 2) if current_h else "—",
            round(diff, 2) if diff is not None else "—",
            rec, "",
        ], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(bold=(col == 7),
                              color=(C_GREEN_FT if rec.startswith("✔") else C_RED_FT) if col == 7 else "000000")
            cell.fill = _fill(C_GREEN_BG if rec.startswith("✔") and col == 7 else
                              (C_RED_BG if col == 7 else (C_ALT_ROW if alt else C_WHITE)))
            cell.alignment = _align("left" if col in (1, 2, 7) else "center", wrap=(col == 2))
            cell.border = _border()
        row += 1

    # Nota metodológica
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1,
        value=(f"Nota: Duração estimada calculada com base na velocidade média real de "
               f"{avg_kbsec / 1024:.0f} MB/s observada nos jobs NetApp deste mês. "
               f"Meta configurável em MAX_NETAPP_HOURS no início do script."))
    cell.font = _font(italic=True, size=9, color="595959")
    cell.fill = _fill(C_WHITE)
    cell.alignment = _align("left", wrap=True)
    ws.row_dimensions[row].height = 30

    _set_col_widths(ws, [22, 55, 18, 22, 22, 16, 30, 10])

# ---------------------------------------------------------------------------
# Aba 6 — Volumes NetApp (crescimento)
# ---------------------------------------------------------------------------

def _sheet_netapp_volumes(wb, volumes):
    ws = wb.create_sheet("NetApp — Volumes")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    dates = sorted(set(v["snapshot_date"] for v in volumes))
    vol_names = sorted(set(v["volume_name"] for v in volumes))

    # Índice: (date, vol_name) -> row
    idx = {(v["snapshot_date"], v["volume_name"]): dict(v) for v in volumes}

    # Headers: Volume | Total | [data1 Used | data2 Used | ...] | Crescimento (GB)
    headers = ["Volume", "Cota (GB)"] + [f"Usado {d}" for d in dates] + ["Cresc. último mês (GB)", "% Cota usada"]
    widths  = [22, 14] + [18] * len(dates) + [24, 18]
    _header_row(ws, 1, headers, widths)

    for i, vol in enumerate(vol_names):
        row = 2 + i
        alt = i % 2 == 1

        total = idx.get((dates[-1], vol), {}).get("total_gb") if dates else None
        used_vals = [idx.get((d, vol), {}).get("used_gb") for d in dates]

        if len(used_vals) >= 2 and used_vals[-1] is not None and used_vals[-2] is not None:
            growth = used_vals[-1] - used_vals[-2]
        else:
            growth = None

        pct_used = (used_vals[-1] / total * 100) if (used_vals and used_vals[-1] and total) else None

        row_data = [vol, round(total, 1) if total else "—"] + \
                   [round(u, 1) if u is not None else "—" for u in used_vals] + \
                   [round(growth, 1) if growth is not None else "—",
                    f"{pct_used:.1f}%" if pct_used else "—"]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font()
            cell.fill = _fill(C_ALT_ROW if alt else C_WHITE)

            # Colorir crescimento
            if col == len(dates) + 3 and growth is not None:
                if growth > 50:
                    cell.fill = _fill(C_RED_BG)
                    cell.font = _font(color=C_RED_FT)
                elif growth > 10:
                    cell.fill = _fill(C_YELLOW_BG)
                    cell.font = _font(color=C_YELLOW_FT)
                elif growth < 0:
                    cell.fill = _fill(C_GREEN_BG)
                    cell.font = _font(color=C_GREEN_FT)

            # Colorir % utilização
            if col == len(dates) + 4 and pct_used is not None:
                if pct_used >= 90:
                    cell.fill = _fill(C_RED_BG)
                    cell.font = _font(color=C_RED_FT)
                elif pct_used >= 75:
                    cell.fill = _fill(C_YELLOW_BG)
                    cell.font = _font(color=C_YELLOW_FT)

            cell.alignment = _align("center" if col > 1 else "left")
            cell.border = _border()

# ---------------------------------------------------------------------------
# Aba 7 — Fitas
# ---------------------------------------------------------------------------

def _sheet_fitas(wb, tapes):
    ws = wb.create_sheet("Fitas — Tapes")
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, "UTILIZAÇÃO DE FITAS LTO POR POOL", 6)
    _header_row(ws, 2, ["Pool", "Qtd Fitas", "Volume Total (TB)", "Volume Médio/Fita (TB)", "Data Snapshot", ""],
                widths=[28, 14, 20, 24, 18, 10])

    # Agrupa por (snapshot_date, pool)
    pools = {}
    for t in tapes:
        key = (t["snapshot_date"], t["pool"])
        if key not in pools:
            pools[key] = {"qty": 0, "kb": 0}
        pools[key]["qty"] += 1
        pools[key]["kb"] += (t["kilobytes"] or 0)

    row = 3
    for i, ((date, pool), data) in enumerate(sorted(pools.items())):
        total_tb = data["kb"] / (1024 ** 3)
        avg_tb   = total_tb / data["qty"] if data["qty"] else 0
        alt = i % 2 == 1
        for col, val in enumerate([pool, data["qty"], round(total_tb, 2), round(avg_tb, 3), date, ""], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font()
            cell.fill = _fill(C_ALT_ROW if alt else C_WHITE)
            cell.alignment = _align("center" if col > 1 else "left")
            cell.border = _border()
        row += 1

    # Totais por data
    row += 1
    _section_title(ws, row, "TOTAIS POR DATA DE SNAPSHOT", 6)
    row += 1
    _header_row(ws, row, ["Data", "Total Fitas", "Volume Total (TB)", "", "", ""])
    row += 1

    by_date = {}
    for t in tapes:
        d = t["snapshot_date"]
        by_date.setdefault(d, {"qty": 0, "kb": 0})
        by_date[d]["qty"] += 1
        by_date[d]["kb"] += (t["kilobytes"] or 0)

    for i, (date, data) in enumerate(sorted(by_date.items())):
        total_tb = data["kb"] / (1024 ** 3)
        alt = i % 2 == 1
        for col, val in enumerate([date, data["qty"], round(total_tb, 2), "", "", ""], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(bold=(col in (1, 2, 3)))
            cell.fill = _fill(C_ALT_ROW if alt else C_WHITE)
            cell.alignment = _align("center" if col > 1 else "left")
            cell.border = _border()
        row += 1

# ---------------------------------------------------------------------------
# Aba 8 — Capacidade
# ---------------------------------------------------------------------------  

def _sheet_capacidade_fitas(wb, tapes, tapes_history, db_path):
    latest_tape_date = None
    """
    Aba Nova — Capacidade de Fitas (Evolução Histórica e Gráfico de Linha)
    Calcula o mês corrente baseado em netbackup_tapes, atualiza o histórico e plota o gráfico.
    """
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    
    ws = wb.create_sheet("Capacidade — Fitas")
    ws.sheet_view.showGridLines = False

# --- 1. PROCESSA E ATUALIZA MÊS ATUAL ---
    # Correção: Acessando diretamente por colchetes, já que sqlite3.Row não aceita .get()
    total_kb_atual = sum((t["kilobytes"] if t["kilobytes"] is not None else 0) for t in tapes if t["snapshot_date"] == latest_tape_date)
    total_tb_atual = total_kb_atual / (1024 ** 3) # KB para TB

    # Identifica o mês do snapshot atual das fitas (Formato: YYYY-MM)
    latest_tape_date = max((t["snapshot_date"] for t in tapes if t["snapshot_date"]), default=None)
    if latest_tape_date:
        try:
            # Tenta tratar formatos comuns como YYYY-MM-DD
            dt_base = datetime.strptime(latest_tape_date[:10], "%Y-%m-%d")        
        except:
            dt_base = datetime.now()
    else:
        dt_base = datetime.now()
    
    dt_mes_anterior = dt_base - pd.Timedelta(days=dt_base.day + 2)
    mes_atual_str = dt_mes_anterior.strftime("%Y-%m")

    # Atualiza em banco de dados para o histórico persistir nas próximas execuções
    if db_path and total_tb_atual > 0:
        conn = sqlite3.connect(db_path)
        try:
            conn.execute(
                "INSERT OR REPLACE INTO netbackup_tapes_history (snapshot_month, used_tb) VALUES (?, ?)",
                (mes_atual_str, round(total_tb_atual, 2))
            )
            conn.commit()
            # Atualiza a lista local para plotagem imediata no relatório atual
            tapes_history = conn.execute("SELECT * FROM netbackup_tapes_history ORDER BY snapshot_month").fetchall()
        except Exception as e:
            print(f"Aviso ao salvar histórico de fitas: {e}")
        finally:
            conn.close()

    # Filtra e captura os últimos 12 registros de meses para não estourar o tamanho horizontal da tela
    historico_12_meses = sorted(
        [{"snapshot_month": row[0], "used_tb": row[1]} for row in tapes_history],
        key=lambda x: x["snapshot_month"]
    )[-12:]

    # --- 2. CONSTRUÇÃO DA TABELA DE CAPACIDADE HISTÓRICA ---
    # Título da Seção
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(historico_12_meses) + 1)
    title_cell = ws.cell(row=1, column=1, value="Capacidade Total de Backup - CPTM em TB")
    title_cell.font = _font(bold=True, size=14, color=C_HEADER_FT)
    title_cell.fill = _fill(C_ACCENT) # Usa o azul vibrante de destaque para bater com a imagem 2
    title_cell.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 28

    # Monta os rótulos horizontais dos meses (convertendo '2025-05' para 'mai/25')
    meses_traducao = {
        "01": "jan", "02": "fev", "03": "mar", "04": "abr", "05": "mai", "06": "jun",
        "07": "jul", "08": "ago", "09": "set", "10": "out", "11": "nov", "12": "dez"
    }
    
    headers_meses = ["Métrica / Período"]
    for row in historico_12_meses:
        ano, mes = row["snapshot_month"].split("-")
        headers_meses.append(f"{meses_traducao.get(mes, mes)}/{ano[2:]}")

    # Escreve Cabeçalho de Meses
    for col_idx, text in enumerate(headers_meses, 1):
        cell = ws.cell(row=2, column=col_idx, value=text)
        cell.font = _font(bold=True, color="000000")
        cell.fill = _fill(C_GRAY_BG if col_idx == 1 else C_WHITE)
        cell.alignment = _align("center", "center")
        cell.border = _border()
    ws.row_dimensions[2].height = 22

    # ALTERAR QUANDO MAIS FITAS FOREM ADICIONADAS - ATUALMENTE CONSIDERANDO FITAS X 12 TB CADA
    total_capacity = 7200.0 # 600 fitas * 12 TB fixos informados
    
    row_capacidade = ["Capacidade Total"] + [total_capacity] * len(historico_12_meses)
    row_em_uso = ["Espaço Backup - Em Uso"] + [r["used_tb"] for r in historico_12_meses]
    row_disponivel = ["Espaço Backup - Disponível"] + [round(total_capacity - r["used_tb"], 2) for r in historico_12_meses]

    for idx_r, dados_linha in enumerate([row_capacidade, row_em_uso, row_disponivel], 3):
        for col_idx, val in enumerate(dados_linha, 1):
            cell = ws.cell(row=idx_r, column=col_idx, value=val)
            cell.border = _border()
            
            if col_idx == 1:
                cell.font = _font(bold=False)
                cell.fill = _fill(C_WHITE)
                cell.alignment = _align("left")
            else:
                cell.font = _font()
                cell.fill = _fill(C_WHITE)
                cell.alignment = _align("center")
                cell.number_format = '#,##0.00'
        ws.row_dimensions[idx_r].height = 20

    # Destaca a primeira coluna de cabeçalhos estáticos de métrica
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(headers_meses) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    # --- 3. ALGORITMO DE PREVISÃO DE ESGOTAMENTO (MÉDIA MÓVEL LINEAR) ---
    row_analise = 7
    _section_title(ws, row_analise, "ANÁLISE PREDITIVA DE ESGOTAMENTO DE CAPACIDADE", 5, bg=C_HEADER_BG, ft=C_HEADER_FT)
    
    # Calcula crescimento médio mensal real
    crescimentos = []
    for idx in range(1, len(historico_12_meses)):
        diff = historico_12_meses[idx]["used_tb"] - historico_12_meses[idx-1]["used_tb"]
        crescimentos.append(diff)
        
    avg_crescimento_mensal = sum(crescimentos) / len(crescimentos) if crescimentos else 0.0
    ultimo_uso_tb = historico_12_meses[-1]["used_tb"]
    espaco_disponivel_tb = total_capacity - ultimo_uso_tb
    
    if avg_crescimento_mensal > 0:
        meses_restantes = espaco_disponivel_tb / avg_crescimento_mensal
        data_esgotamento = datetime.now() + pd.Timedelta(days=int(meses_restantes * 30.4)) # type: ignore
        status_previsao = f"Esgotamento previsto em aproximadamente {int(meses_restantes)} meses ({data_esgotamento.strftime('%m/%Y')})."
        previsao_color_bg = C_YELLOW_BG
        previsao_color_ft = C_YELLOW_FT
        if meses_restantes <= 6: # Crítico se faltar menos de 6 meses
            previsao_color_bg = C_RED_BG
            previsao_color_ft = C_RED_FT
    else:
        status_previsao = "✔ Estabilizado: Crescimento nulo ou negativo observado nos últimos meses."
        previsao_color_bg = C_GREEN_BG
        previsao_color_ft = C_GREEN_FT

    _header_row(ws, row_analise + 1, ["Métrica Preditiva", "Valor Calculado"], widths=[32, 55])
    
    indicadores = [
        ("Crescimento Médio Mensal Observado", f"{avg_crescimento_mensal:.2f} TB / mês"),
        ("Espaço Livre em Estoque (Fitas)", f"{espaco_disponivel_tb:.2f} TB"),
        ("Status de Atenção da Janela", status_previsao)
    ]
    
    row_curr = row_analise + 2
    for i, (lab, val) in enumerate(indicadores):
        c1 = ws.cell(row=row_curr, column=1, value=lab)
        c2 = ws.cell(row=row_curr, column=2, value=val)
        c1.border = _border(); c2.border = _border()
        c1.font = _font(); c1.fill = _fill(C_ALT_ROW if i % 2 == 0 else C_WHITE)
        
        if i == 2: # Destaca a linha de previsão
            c2.font = _font(bold=True, color=previsao_color_ft)
            c2.fill = _fill(previsao_color_bg)
        else:
            c2.font = _font(bold=True)
            c2.fill = _fill(C_ALT_ROW if i % 2 == 0 else C_WHITE)
        row_curr += 1


    # --- 4. GERAÇÃO DO GRÁFICO DE LINHAS (PADRÃO CPTM EXECUTIVO) ---
    chart = LineChart()
    chart.title = "Volume Backup CPTM - Total X Em Uso (TB) - Mês"
    chart.style = 13 # Estilo nativo limpo do Excel
    chart.y_axis.title = "Capacidade (TB)"
    chart.x_axis.title = "Meses de Referência"
    chart.width = 24
    chart.height = 13
    
    # Referência dos dados (Linhas 3 a 5 da planilha, cobrindo as colunas dos meses)
    data_ref = Reference(ws, min_col=1, min_row=3, max_col=len(historico_12_meses) + 1, max_row=5)
    # Referência das categorias do eixo X (Linha 2, colunas dos meses)
    cats_ref = Reference(ws, min_col=2, min_row=2, max_col=len(historico_12_meses) + 1, max_row=2)
    
    chart.add_data(data_ref, titles_from_data=True, from_rows=True)
    chart.set_categories(cats_ref)
    
    # Habilita rótulos de dados flutuantes sobre a linha (Data Labels) para simular o seu print
    for serie in chart.series:
        serie.dataLabels = DataLabelList()
        serie.dataLabels.showVal = True
        serie.graphicalProperties.line.width = 25000  # Engrossa um pouco a linha para melhor visibilidade
        
    # Customização manual das cores das linhas no gráfico para bater com a imagem informada
    if len(chart.series) >= 3:
        chart.series[0].graphicalProperties.solidFill = "2E75B6" # Linha Total -> Azul Destaque
        chart.series[1].graphicalProperties.solidFill = "E67E22" # Linha Em Uso -> Laranja
        chart.series[2].graphicalProperties.solidFill = "A6ACAF" # Linha Disponível -> Cinza

    # Posiciona o gráfico logo abaixo dos quadros de análise preditiva
    ws.add_chart(chart, "A13")      

#----------------------------------------------------------------------------
# Aba 9 — NetApp — 15 maiores volumes por uso
#----------------------------------------------------------------------------

def _sheet_netapp_15_maiores(wb, volumes):
    """
    Aba Nova — Análise dos 15 Maiores Volumes NetApp (Percentual, Tamanho Absoluto e Crescimento)
    Gera as três matrizes analíticas solicitadas e plota o gráfico de barras empilhadas de cota.
    """
    from openpyxl.chart import BarChart, Reference
    
    ws = wb.create_sheet("15 Maiores — NetApp")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # --- 1. PROCESSAMENTO DE DATAS (ATUAL X ANTERIOR) ---
    todas_datas = sorted(list(set(v["snapshot_date"] for v in volumes)))
    if len(todas_datas) < 1:
        return
        
    data_atual = todas_datas[-1]
    data_anterior = todas_datas[-2] if len(todas_datas) >= 2 else None

    # Agrupa dados em dicionários indexados pelo nome do volume para cálculos ágeis
    vols_atuais = {v["volume_name"].upper(): dict(v) for v in volumes if v["snapshot_date"] == data_atual}
    vols_anteriores = {v["volume_name"].upper(): dict(v) for v in volumes if v["snapshot_date"] == data_anterior} if data_anterior else {}

    # --- 2. GERAÇÃO DAS REGRAS DE NEGÓCIO DE TOP 15 ---
    lista_analitica = []
    for name, v in vols_atuais.items():
        total = v["total_gb"] or 0.0
        used = v["used_gb"] or 0.0
        free = v["free_gb"] or 0.0
        pct = (used / total * 100) if total > 0 else 0.0
        
        # Calcula crescimento em relação ao snapshot imediatamente anterior
        v_ant = vols_anteriores.get(name)
        growth = (used - v_ant["used_gb"]) if v_ant and v_ant["used_gb"] is not None else 0.0
        
        lista_analitica.append({
            "name": name, "total": total, "used": used, "free": free, "pct": pct, "growth": growth
        })

    # Extrações de Rankings (Top 15 de cada métrica)
    top15_pct = sorted(lista_analitica, key=lambda x: x["pct"], reverse=True)[:15]
    top15_used = sorted(lista_analitica, key=lambda x: x["used"], reverse=True)[:15]
    top15_growth = sorted(lista_analitica, key=lambda x: x["growth"], reverse=True)[:15]

    # --- 3. CONSTRUÇÃO VISUAL DA TABELA 1 (MAIORES POR PERCENTUAL) ---
    # Título Principal da Seção Esquerda
    ws.merge_cells("A1:E1")
    t1 = ws["A1"]
    t1.value = "Capacidade Cota - CPTM 15 Maiores Utilizados (Por Percentual)"
    t1.font = _font(bold=True, size=11, color=C_HEADER_FT)
    t1.fill = _fill("262626") # Fundo escuro conforme imagem 4
    t1.alignment = _align("center")
    ws.row_dimensions[1].height = 24

    headers_t1 = ["Pasta", "Cota Total da Pasta (GB)", "Usado GB", "Cota em Uso (%)", "Disponível GB"]
    _header_row(ws, 2, headers_t1, bg=C_GRAY_BG, ft="000000")
    ws.row_dimensions[2].height = 20

    for idx, item in enumerate(top15_pct):
        r = 3 + idx
        alt = idx % 2 == 1
        # Injeta prefixo fictício '\\PARIS\' para bater exatamente com a estética operacional da imagem 4
        vals = [f"\\\\PARIS\\{item['name']}", item["total"], item["used"], item["pct"]/100, item["free"]]
        _data_row(ws, r, vals, alt=alt)
        
        # Formatações numéricas customizadas
        ws.cell(row=r, column=2).number_format = '#,##0'
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4).number_format = '0.00%'
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=1).alignment = _align("left")
        for col in range(2, 6):
            ws.cell(row=r, column=col).alignment = _align("right")
        ws.row_dimensions[r].height = 19

    # --- 4. CONSTRUÇÃO VISUAL DA TABELA 2 (MAIORES POR TAMANHO ABSOLUTO) ---
    # Posicionada ao lado direito (Colunas G a K)
    ws.merge_cells("G1:K1")
    t2 = ws["G1"]
    t2.value = "Volume Absoluto - CPTM 15 Maiores Espaços Ocupados (GB)"
    t2.font = _font(bold=True, size=11, color=C_HEADER_FT)
    t2.fill = _fill(C_HEADER_BG)
    t2.alignment = _align("center")

    # Escreve o cabeçalho da Tabela 2
    for col_idx, text in enumerate(headers_t1, 7):
        cell = ws.cell(row=2, column=col_idx, value=text)
        cell.font = _font(bold=True, color="000000")
        cell.fill = _fill(C_GRAY_BG)
        cell.alignment = _align("center")
        cell.border = _border()

    for idx, item in enumerate(top15_used):
        r = 3 + idx
        alt = idx % 2 == 1
        vals = [f"\\\\PARIS\\{item['name']}", item["total"], item["used"], item["pct"]/100, item["free"]]
        
        for c_offset, val in enumerate(vals, 7):
            cell = ws.cell(row=r, column=c_offset, value=val)
            cell.font = _font()
            cell.fill = _fill(C_ALT_ROW if alt else C_WHITE)
            cell.border = _border()
            
            if c_offset == 7:
                cell.alignment = _align("left")
            else:
                cell.alignment = _align("right")
                if c_offset == 10:
                    cell.number_format = '0.00%'
                elif c_offset == 8:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.00'

    # --- 5. CONSTRUÇÃO VISUAL DA TABELA 3 (MAIORES TAXAS DE CRESCIMENTO) ---
    # Posicionada abaixo da tabela 2 (Coluna G, linha 21)
    start_r_t3 = 21
    ws.merge_cells(start_row=start_r_t3, start_column=7, end_row=start_r_t3, end_column=10)
    t3 = ws.cell(row=start_r_t3, column=7)
    t3.value = f"Impacto Mensal — 15 Maiores Crescimentos (Base: {data_anterior or 'Mês Ant.'} X {data_atual})"
    t3.font = _font(bold=True, size=11, color=C_HEADER_FT)
    t3.fill = _fill(C_ORANGE_BG)
    ws.cell(row=start_r_t3, column=7).font = _font(bold=True, size=11, color=C_ORANGE_FT)
    t3.alignment = _align("center")
    ws.row_dimensions[start_r_t3].height = 24

    headers_t3 = ["Pasta Target", "Ocupação Anterior (GB)", "Ocupação Atual (GB)", "Variação Mensal Delta (GB)"]
    for col_idx, text in enumerate(headers_t3, 7):
        cell = ws.cell(row=start_r_t3 + 1, column=col_idx, value=text)
        cell.font = _font(bold=True, color=C_HEADER_FT)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = _align("center")
        cell.border = _border()
    ws.row_dimensions[start_r_t3 + 1].height = 20

    for idx, item in enumerate(top15_growth):
        r = start_r_t3 + 2 + idx
        alt = idx % 2 == 1
        vol_ant_used = vols_anteriores.get(item["name"], {}).get("used_gb", 0.0)
        
        vals = [f"\\\\PARIS\\{item['name']}", vol_ant_used, item["used"], item["growth"]]
        
        for c_offset, val in enumerate(vals, 7):
            cell = ws.cell(row=r, column=c_offset, value=val)
            cell.font = _font()
            cell.fill = _fill(C_ALT_ROW if alt else C_WHITE)
            cell.border = _border()
            
            if c_offset == 7:
                cell.alignment = _align("left")
            else:
                cell.alignment = _align("right")
                cell.number_format = '#,##0.00'
                
            # Destaque condicional sutil: se cresceu mais de 50GB em um único mês, pinta a célula de variação de vermelho
            if c_offset == 10 and item["growth"] > 50.0:
                cell.fill = _fill(C_RED_BG)
                cell.font = _font(bold=True, color=C_RED_FT)
        ws.row_dimensions[r].height = 19

    # Ajuste milimétrico das larguras das colunas para evitar o estouro de texto (###)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 4   # Coluna de espaçamento em branco entre as tabelas
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 24
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 15

    # --- 6. IMPLEMENTAÇÃO DO GRÁFICO DE BARRAS EMPILHADAS (100% COTA) ---
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "percentStacked" # Força o empilhamento proporcional a 100% igual ao print 5
    chart.overlap = 100
    chart.title = "15 Maiores Usos (%)"
    chart.width = 19
    chart.height = 14
    
    # Captura dados das colunas Usado (Col C) e Disponível (Col E).
    # O openpyxl precisa ler colunas contíguas ou tratadas. Para o gráfico empilhado ideal de 2 séries:
    # Usaremos referências dinâmicas das colunas de dados criadas na Tabela 1.
    data_ref = Reference(ws, min_col=3, min_row=2, max_col=5, max_row=17) # Pega coluna C até E (Linhas 2 a 17)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=17) # Nomes das Pastas na coluna A
    
    chart.add_data(data_ref, titles_from_data=True, from_rows=False)
    chart.set_categories(cats_ref)
    
    # Remove a série central do percentual (Coluna D) para não duplicar dados no gráfico, 
    # mantendo puramente as parcelas absolutas de Usado e Disponível que formam os 100% da cota.
    if len(chart.series) == 3:
        del chart.series[1] 

    # Configuração de Cores das Séries para replicar perfeitamente a Imagem 5
    chart.series[0].graphicalProperties.solidFill = "F4A460" # Usado GB -> Cor Laranja Suave / Areia
    chart.series[1].graphicalProperties.solidFill = "B0BF1A" # Disponível GB -> Cinza / Verde Oliva Claro
    
    # Força a exibição da legenda na base inferior do gráfico
    chart.legend.position = "b" # type: ignore

    # Insere o gráfico logo abaixo da primeira tabela (Linha 21 da coluna A)
    ws.add_chart(chart, "A21")

#-------------------------------------------------------------------------------
# Aba 10 — Gantt de Jobs
#-------------------------------------------------------------------------------

def _sheet_gantt_jobs(wb, jobs):
    """
    Aba Nova — Diagrama de Gantt Visual dos Jobs de Backup
    Mapeia a linha do tempo dos jobs usando um gráfico de barras horizontais empilhadas.
    """
    from openpyxl.chart import BarChart, Reference
    
    ws = wb.create_sheet("Gantt — Janela Mensal")
    ws.sheet_view.showGridLines = False
    
    # Título Principal da Aba
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "DIAGRAMA DE GANTT — CRONOGRAMA DE EXECUÇÃO DOS JOBS"
    title_cell.font = _font(bold=True, size=11, color=C_HEADER_FT)
    title_cell.fill = _fill(C_HEADER_BG)
    title_cell.alignment = _align("left", "center")
    ws.row_dimensions[1].height = 26

    # Cabeçalho da Tabela de Dados de Suporte do Gráfico
    headers = ["Política / Job", "Data/Hora Início", "Ponto de Partida (Excel Num)", "Duração Real (Dias)"]
    _header_row(ws, 2, headers, widths=[35, 18, 22, 18])
    ws.row_dimensions[2].height = 20

    # 1. FILTRAGEM E ORDENAÇÃO DOS JOBS
    valid_jobs = []
    for j in jobs:
        try:
            job_dict = dict(j)
        except:
            job_dict = j
            
        start_time = job_dict.get("start_time")
        elapsed_time = job_dict.get("elapsed_time")
        policy = job_dict.get("job_policy") or "Sem Política"
        job_id = job_dict.get("job_id")
        
        if not start_time or not elapsed_time:
            continue
            
        dt_start = _parse_dt(start_time)
        dur_hours = _iso_to_hours(elapsed_time)
        
        # Filtro: Foca apenas nos jobs relevantes da janela mensal (descarta ruídos menores de 15 min)
        if dur_hours < 0.25:
            continue
            
        valid_jobs.append({
            "label": f"{policy} (ID: {job_id})",
            "dt_start": dt_start,
            "dur_days": dur_hours / 24.0 # Transforma horas em fração de dia para o Excel calcular o Gantt
        })
        
    # Ordena os jobs cronologicamente pelo horário de início (fundamental para o Gantt fazer sentido)
    valid_jobs = sorted(valid_jobs, key=lambda x: x["dt_start"])
    
    # Limita aos 35 maiores/principais se a lista for gigantesca, para o gráfico não virar um bloco preto ilegível
    valid_jobs = valid_jobs[:35]

    if not valid_jobs:
        ws.cell(row=3, column=1, value="Nenhum job pesado mapeado para esta janela mensal.")
        return

    # Descobre o ponto zero do gráfico (o primeiro backup a começar no mês)
    primeiro_start_dt = valid_jobs[0]["dt_start"]

    # 2. ESCRITA DOS DADOS NA PLANILHA
    row_idx = 3
    for job in valid_jobs:
        ws.cell(row=row_idx, column=1, value=job["label"])
        ws.cell(row=row_idx, column=2, value=job["dt_start"].astimezone().strftime("%d/%m %H:%M"))
        
        # Ponto de Partida: Diferença em dias desde o primeiro job para criar o deslocamento horizontal
        delta_dias = (job["dt_start"] - primeiro_start_dt).total_seconds() / 86400.0
        ws.cell(row=row_idx, column=3, value=delta_dias)
        
        # Duração real em fração de dias
        ws.cell(row=row_idx, column=4, value=job["dur_days"])
        
        # Formatações Visuais de suporte
        bg_row = C_ALT_ROW if row_idx % 2 == 0 else C_WHITE
        for col in range(1, 5):
            c = ws.cell(row=row_idx, column=col)
            c.fill = _fill(bg_row)
            c.font = _font()
            c.border = _border()
            c.alignment = _align("center" if col > 1 else "left")
            if col in [3, 4]:
                c.number_format = '0.0000'
                
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    # --- 3. CONSTRUÇÃO E RENDEREZAÇÃO DO GRÁFICO DE GANTT ---
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"  # Empilhado: Essencial para o efeito de Gantt
    chart.overlap = 100
    chart.title = f"Linha do Tempo e Duração dos Backups Mensais - Referência: {primeiro_start_dt.strftime('%m/%Y')}"
    chart.width = 25
    chart.height = 16
    
    # Eixos do gráfico
    chart.x_axis.title = "Políticas de Backup Executadas"
    chart.y_axis.title = f"Linha do Tempo (Dias Corridos a partir de {primeiro_start_dt.strftime('%d/%m %H:%M')})"
    
    # Referências do Excel (Dados das colunas C e D | Categorias da coluna A)
    data_ref = Reference(ws, min_col=3, min_row=2, max_col=4, max_row=row_idx-1)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=row_idx-1)
    
    chart.add_data(data_ref, titles_from_data=True, from_rows=False)
    chart.set_categories(cats_ref)
    
    # O truque de mágica do Gantt no Excel:
    # A Série 1 (Ponto de Partida/Deslocamento) DEVE ficar invisível.
    # A Série 2 (Duração Real) assume a cor de preenchimento de destaque da CPTM.
    if len(chart.series) >= 2:
        # Torna a barra de deslocamento invisível tirando as propriedades gráficas dela
        chart.series[0].graphicalProperties.solidFill = None
        chart.series[0].graphicalProperties.line.solidFill = None
        
        # Dá o destaque azul corporativo para a barra que representa a duração real do backup
        chart.series[1].graphicalProperties.solidFill = C_HEADER_BG
        chart.series[1].graphicalProperties.line.solidFill = C_HEADER_BG

    # Esconde a legenda para o gráfico ficar limpo, já que o Gantt é autoexplicativo
    chart.legend = None # type: ignore

    # Posiciona o Gráfico de Gantt de forma imponente ao lado direito da tabela auxiliar
    ws.add_chart(chart, "F2")

# -------------------------------------------------------------------------------
# Graficos das Janelas
# -------------------------------------------------------------------------------

def _sheet_graficos_janelas(wb, jobs):
    """
    Aba Nova — Agrupamento e Visão Gráfica de Tempo Total por Tipo de Política.
    Gera 4 gráficos de barras horizontais replicando o modelo visual solicitado.
    """
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    
    ws = wb.create_sheet("Janelas — Visão Gráfica")
    ws.sheet_view.showGridLines = False
    
    # -------------------------------------------------------------------------
    # 1. CLASSIFICAÇÃO E TRATAMENTO DA REGRA DE NEGÓCIO (PROCESSO DE DADOS)
    # -------------------------------------------------------------------------
    # Dicionários de agrupamento para reter o job mais longo { nome_politica: max_horas }
    dados_categorias = {
        "NetApp": {},
        "Oracle": {},
        "VMWares": {},
        "Client Windows": {}
    }

    for j in jobs:
        try:
            job_dict = dict(j)
        except:
            job_dict = j
            
        policy_name = job_dict.get("job_policy")
        elapsed_str = job_dict.get("elapsed_time")
        
        if not policy_name or not elapsed_str:
            continue
            
        policy_upper = policy_name.upper()
        horas = _iso_to_hours(elapsed_str) # Converte a duração ISO para float
        
        # Define a categoria baseado nos prefixos e regras da CPTM
        if policy_upper.startswith("NETAPP_"):
            cat = "NetApp"
        elif policy_upper.startswith("ORACLE_"):
            cat = "Oracle"
        elif policy_upper.startswith("VMWARES_"):
            cat = "VMWares"
        else:
            cat = "Client Windows"
            
        # Regra de Ouro (Aplica para todas, garantindo o teto/máximo para VMWares):
        # Se a política já existir, mantém apenas o registro que levou mais tempo.
        if policy_name not in dados_categorias[cat] or horas > dados_categorias[cat][policy_name]:
            dados_categorias[cat][policy_name] = horas

    # -------------------------------------------------------------------------
    # 2. ESCRITA DAS MATRIZES DE SUPORTE NA PLANILHA (COLUNAS A até C)
    # -------------------------------------------------------------------------
    # Cabeçalho unificado para a tabela auxiliar de dados
    _header_row(ws, 1, ["Categoria", "Política / Client", "Tempo Total (Horas)"], widths=[16, 25, 20])
    
    # Dicionário para guardar as coordenadas de início e fim de cada bloco para alimentar os gráficos
    coordenadas_mapa = {}
    row_atual = 2

    for cat, pol_dict in dados_categorias.items():
        if not pol_dict:
            continue
            
        start_row = row_atual
        # Ordena as políticas pelo nome de forma ascendente para manter o padrão sequencial (Ex: VMWares_1, 2, 3...)
        for pol_name, horas in sorted(pol_dict.items()):
            ws.cell(row=row_atual, column=1, value=cat)
            ws.cell(row=row_atual, column=2, value=pol_name)
            ws.cell(row=row_atual, column=3, value=round(horas, 2))
            
            # Formatação sutil da linha de dados
            alt = (row_atual % 2 == 1)
            for col in range(1, 4):
                cell = ws.cell(row=row_atual, column=col)
                cell.font = _font()
                cell.fill = _fill(C_ALT_ROW if alt else C_WHITE)
                cell.border = _border()
                if col == 3:
                    cell.alignment = _align("right")
                    cell.number_format = '#,##0.00"h"'
                else:
                    cell.alignment = _align("left")
            row_atual += 1
            
        end_row = row_atual - 1
        coordenadas_mapa[cat] = (start_row, end_row)

    # -------------------------------------------------------------------------
    # 3. CONSTRUÇÃO E ALINHAMENTO DOS 4 GRÁFICOS (GRID VISUAL 2X2)
    # -------------------------------------------------------------------------
    # Mapa de posições das células onde os gráficos vão aterrissar na tela
    posicoes_graficos = {
        "NetApp": "E2",          # Superior Esquerdo
        "Oracle": "N2",          # Superior Direito
        "VMWares": "E20",        # Inferior Esquerdo
        "Client Windows": "N20"  # Inferior Direito
    }

    for cat, local_celula in posicoes_graficos.items():
        if cat not in coordenadas_mapa:
            continue
            
        start_r, end_r = coordenadas_mapa[cat]
        
        # Instanciação do Gráfico de Barras Horizontais conforme modelo enviado
        chart = BarChart()
        chart.type = "bar"         # "bar" horizontal (se fosse vertical seria "col")
        chart.style = 10           # Estilo nativo limpo
        chart.title = f"Tempo Total de Backup - {cat}"
        chart.width = 16
        chart.height = 11
        
        # Referência dos Dados (Coluna C: Tempo Total em Horas)
        data_ref = Reference(ws, min_col=3, min_row=start_r, max_row=end_r)
        # Referência das Categorias/Eixo Y (Coluna B: Nome da Política)
        cats_ref = Reference(ws, min_col=2, min_row=start_r, max_row=end_r)
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        
        # Configurações do Título do Eixo X (Duração na horizontal)
        chart.x_axis.title = "Duração (Horas)"
        chart.y_axis.title = "Políticas"
        
        # Costumização das Cores das Barras para o Azul Padrão da CPTM (C_ACCENT)
        if len(chart.series) > 0:
            chart.series[0].graphicalProperties.solidFill = C_ACCENT
            chart.series[0].graphicalProperties.line.solidFill = C_ACCENT
            # Adiciona os rótulos numéricos de tempo em cima de cada barra para facilitar leitura imediata
            chart.series[0].dataLabels = DataLabelList()
            chart.series[0].dataLabels.showVal = True

        # Remove a legenda lateral desnecessária (já que o título já diz a que se refere)
        chart.legend = None # type: ignore
        
        # Renderiza o gráfico na posição correta do grid
        ws.add_chart(chart, local_celula)

    # Adiciona a importação necessária do DataLabelList localmente caso não mapeada no escopo global
    import openpyxl.chart.label

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate_report(db_path=DB_PATH, output_dir=OUTPUT_DIR):
    print(f"Lendo banco: {db_path}")
    avail, jobs, pols, tapes, volumes, tapes_history = _load(db_path)

    # Detecta mês de referência a partir dos jobs
    dates_in_jobs = [_parse_dt(j["start_time"]) for j in jobs if j["start_time"]]
    if dates_in_jobs:
        ref_dt = min(dates_in_jobs)
        report_month = ref_dt.strftime("%B/%Y").capitalize()
    else:
        report_month = datetime.now().strftime("%B/%Y").capitalize()

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão

    print("Gerando abas...")
    _sheet_capa(wb, avail, jobs, volumes, report_month)
    _sheet_disponibilidade(wb, avail)
    _sheet_backup_geral(wb, jobs)
    _sheet_janelas(wb, jobs)
    _sheet_netapp_split(wb, jobs, pols, volumes)
    _sheet_netapp_volumes(wb, volumes)
    _sheet_fitas(wb, tapes)
    _sheet_capacidade_fitas(wb, tapes, tapes_history, db_path)
    _sheet_netapp_15_maiores(wb, volumes)
    _sheet_gantt_jobs(wb, jobs)
    _sheet_graficos_janelas(wb, jobs)

    # Salva com timestamp
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = output_dir / f"relatorio_ti_cptm_{ts}.xlsx"
    wb.save(filename)
    print(f"Relatório gerado: {filename}")
    return filename


if __name__ == "__main__":
    generate_report()